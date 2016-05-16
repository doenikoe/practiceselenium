from behave import *
from openpyxl import load_workbook

@given('I open practiceselenium website')
def step_impl(context):
    context.browser.get("http://practiceselenium.com")

@when('I click on menu "{menu}"')
def step_impl(context, menu):
    context.browser.find_element_by_link_text(menu).click()

@then('"{menu}" page link is correct')
def testing(context, menu):
    wb = load_workbook(filename='navigation.xlsx')
    ws = wb.active
    rownum = 0
    found = 0
    for row in ws.iter_rows():
        rownum += 1
        for cell in row:
            if menu == cell.value:
                found = 1
                assert context.browser.current_url in ws.cell(row=rownum, column=2).value

    if 0 == found:
        raise Exception('data not found in fixtures')
